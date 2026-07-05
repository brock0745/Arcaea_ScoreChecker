#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Arcaea 理論値チェッカー - songs.json 更新スクリプト

使い方:
  # コンサルタントシート(xlsx)から曲名・ノーツ数・定数を取り込んでマージ
  python3 update.py --xlsx "コンサルタントシート.xlsx"

  # 単曲を手動で追加/更新
  python3 update.py --add --title "Testify" --disp "Testify" --diff BYD --tier 38 --notes 2221 --cc 12.0 --jacket j0567.webp

songs.json の1エントリの形式:
  {
    "title": "英題（ゲーム内表記）",
    "disp":  "表示名（和題があれば和題、なければ英題と同じ）",
    "diff":  "FTR" | "BYD" | "ETR",
    "tier":  理論値難易度 (0 = 未確定),
    "notes": ノーツ数 (int or null),
    "cc":    譜面定数 (float or null),
    "jacket": "jackets/ 配下のファイル名 (or null)"
  }

このスクリプトは songs.json を「タイトル+難易度」をキーに安全にマージします。
既存曲は notes/cc/jacket が新しい値で更新され、tier は指定があれば上書きされます。
新曲は tier=0（未確定）として追加され、後で理論値難易度表の値が分かり次第
--set-tier で更新するか、songs.json を直接編集してください。
"""
import argparse, json, re, sys, unicodedata, os

SONGS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "songs.json")

def norm_title(s):
    return unicodedata.normalize("NFKC", str(s)).strip().lower().replace("  ", " ")

def load_songs():
    if os.path.exists(SONGS_PATH):
        with open(SONGS_PATH, encoding="utf-8") as f:
            return json.load(f)
    return []

def save_songs(songs):
    with open(SONGS_PATH, "w", encoding="utf-8") as f:
        json.dump(songs, f, ensure_ascii=False, separators=(",", ":"))
    print(f"songs.json 更新完了: {len(songs)} 曲 -> {SONGS_PATH}")

def key_of(title, diff):
    return norm_title(title) + "|" + diff

def merge_song(songs, index, entry):
    """タイトル+難易度でマージ。既存なら更新、無ければ追加。"""
    k = key_of(entry["title"], entry["diff"])
    if k in index:
        i = index[k]
        old = songs[i]
        for field in ("notes", "cc", "jacket", "disp"):
            if entry.get(field) is not None:
                old[field] = entry[field]
        if entry.get("tier") is not None:
            old["tier"] = entry["tier"]
        return "updated"
    else:
        songs.append({
            "title": entry["title"],
            "disp": entry.get("disp") or entry["title"],
            "diff": entry["diff"],
            "tier": entry.get("tier", 0),
            "notes": entry.get("notes"),
            "cc": entry.get("cc"),
            "jacket": entry.get("jacket"),
        })
        index[k] = len(songs) - 1
        return "added"

def build_index(songs):
    return {key_of(s["title"], s["diff"]): i for i, s in enumerate(songs)}

def import_from_xlsx(xlsx_path):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が必要です: pip install openpyxl --break-system-packages")
        sys.exit(1)

    songs = load_songs()
    index = build_index(songs)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "DataM" not in wb.sheetnames:
        print("警告: 'DataM' シートが見つかりません。想定と異なる形式の可能性があります。")
        sys.exit(1)
    ws = wb["DataM"]

    # DataM 列想定: col5(F)=定数(BP), col6(G)=タイトル [DIFF], col8(I)=notes
    DIFF_RE = re.compile(r"^(.*)\s\[(FTR|BYD|ETR)\]$")
    added = updated = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 9:
            continue
        cc_raw, title_raw, notes_raw = row[5], row[6], row[8]
        if not title_raw:
            continue
        m = DIFF_RE.match(str(title_raw).strip())
        if not m:
            continue
        title, diff = m.group(1), m.group(2)
        try:
            cc = float(cc_raw) if cc_raw is not None else None
        except (TypeError, ValueError):
            cc = None
        try:
            notes = int(notes_raw) if notes_raw is not None else None
        except (TypeError, ValueError):
            notes = None
        result = merge_song(songs, index, {"title": title, "diff": diff, "cc": cc, "notes": notes})
        if result == "added":
            added += 1
        else:
            updated += 1

    save_songs(songs)
    print(f"xlsx取込: 新規{added}件 / 更新{updated}件")

def add_single(args):
    songs = load_songs()
    index = build_index(songs)
    entry = {
        "title": args.title,
        "disp": args.disp,
        "diff": args.diff,
        "tier": args.tier,
        "notes": args.notes,
        "cc": args.cc,
        "jacket": args.jacket,
    }
    result = merge_song(songs, index, entry)
    save_songs(songs)
    print(f"{result}: {args.title} [{args.diff}]")

def set_tier(args):
    songs = load_songs()
    index = build_index(songs)
    k = key_of(args.title, args.diff)
    if k not in index:
        print(f"見つかりません: {args.title} [{args.diff}]")
        sys.exit(1)
    songs[index[k]]["tier"] = args.tier
    save_songs(songs)
    print(f"tier更新: {args.title} [{args.diff}] -> {args.tier}")

def main():
    p = argparse.ArgumentParser(description="Arcaea songs.json 更新スクリプト")
    p.add_argument("--xlsx", help="コンサルタントシート(xlsx)のパス。DataMシートから一括更新します。")
    p.add_argument("--add", action="store_true", help="単曲を追加/更新します。")
    p.add_argument("--set-tier", action="store_true", help="既存曲のtier（理論値難易度）だけを更新します。")
    p.add_argument("--title", help="英題（ゲーム内表記）")
    p.add_argument("--disp", help="和題（表示名）。省略時は英題と同じ。")
    p.add_argument("--diff", choices=["FTR", "BYD", "ETR"], help="難易度")
    p.add_argument("--tier", type=int, help="理論値難易度 (0=未確定)")
    p.add_argument("--notes", type=int, help="ノーツ数")
    p.add_argument("--cc", type=float, help="譜面定数")
    p.add_argument("--jacket", help="jackets/ 配下のファイル名")
    args = p.parse_args()

    if args.xlsx:
        import_from_xlsx(args.xlsx)
    elif args.add:
        if not (args.title and args.diff):
            print("--add には --title と --diff が必須です。")
            sys.exit(1)
        add_single(args)
    elif args.set_tier:
        if not (args.title and args.diff and args.tier is not None):
            print("--set-tier には --title --diff --tier が必須です。")
            sys.exit(1)
        set_tier(args)
    else:
        p.print_help()

if __name__ == "__main__":
    main()
